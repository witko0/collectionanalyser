#!/usr/bin/env python3

# configuration
printMissingCards = True

import sys, pprint
# make sure your python libs are in sys path
#sys.path.append('C:\Program Files (x86)\Python37-32\Lib\site-packages')

from openpyxl import load_workbook
wb = load_workbook('lotrtcg_sets.xlsx')

with open('lotrtcg_sets.txt', mode='wt', encoding='utf-8') as myfile:
	myfile.write('Expansions list:\n')
	myfile.write('\n'.join(wb.sheetnames))

for expansion in wb.sheetnames:
	set = wb[expansion]

	cards_no = 0
	cards_C_no = 0
	cards_U_no = 0
	cards_R_no = 0
	cards_P_no = 0
	foils_no = 0
	missing_C = 0
	missing_C_log = []
	missing_U = 0
	missing_U_log = []
	missing_R = 0
	missing_R_log = []
	missing_P = 0
	missing_P_log = []
	missing_Rplus = 0
	missing_Rplus_log = []
	missing_S = 0
	missing_S_log = []
	missing_RF = 0
	missing_RF_log = []
	set_C_completion = 0
	set_U_completion = 0
	set_R_completion = 0
	set_P_completion = 0

	for row in set.rows:
		if row[0].value == 'ID':
			continue
		if row[0].value == None:
			break
		if not row[4].value:
			if 'C' in row[0].value:
				missing_C = missing_C + 1
				missing_C_log.append(row[0].value + ' ' + row[1].value)
			if 'U' in row[0].value:
				missing_U = missing_U + 1
				missing_U_log.append(row[0].value + ' ' + row[1].value)
			if 'R+' in row[0].value:
				missing_Rplus = missing_Rplus + 1
				missing_Rplus_log.append(row[0].value + ' ' + row[1].value)
			if 'R' in row[0].value:
				missing_R = missing_R + 1
				missing_R_log.append(row[0].value + ' ' + row[1].value)
			if 'RF' in row[0].value:
				missing_RF = missing_RF + 1
				missing_RF_log.append(row[0].value + ' ' + row[1].value)
			if 'P' in row[0].value:
				missing_P = missing_P + 1
				missing_P_log.append(row[0].value + ' ' + row[1].value)
			if 'S' in row[0].value:
				missing_S = missing_S + 1
				missing_S_log.append(row[0].value + ' ' + row[1].value)
		if row[5].value:
			foils_no = foils_no + 1
		if row[0].value:
			cards_no = cards_no + 1
			if 'C' in row[0].value:
				cards_C_no = cards_C_no + 1
			if 'U' in row[0].value:
				cards_U_no = cards_U_no + 1
			if 'R' in row[0].value:
				cards_R_no = cards_R_no + 1
			if 'P' in row[0].value:
				cards_P_no = cards_P_no + 1

	missing = missing_C + missing_U + missing_R + missing_P + missing_Rplus + missing_S + missing_RF
	set_completion = 100 - (missing / cards_no) * 100
	if cards_C_no:
		set_C_completion = 100 - (missing_C / cards_C_no) * 100
	if cards_U_no:
		set_U_completion = 100 - (missing_U / cards_U_no) * 100
	if cards_R_no:
		set_R_completion = 100 - (missing_R / cards_R_no) * 100
	if cards_P_no:
		set_P_completion = 100 - (missing_P / cards_P_no) * 100

	with open('lotrtcg_sets.txt', mode='a', encoding='utf-8') as myfile:
		myfile.write('\n\n============================== ' + expansion + ' ==============================\n\n')
		myfile.write('Set completion: ' + str(round(set_completion, 1)) + '%' + '\n')
		if cards_C_no:
			myfile.write('C set completion: ' + str(round(set_C_completion, 1)) + '%' + '\n')
		if cards_U_no:
			myfile.write('U set completion: ' + str(round(set_U_completion, 1)) + '%' + '\n')
		if cards_R_no:
			myfile.write('R set completion: ' + str(round(set_R_completion, 1)) + '%' + '\n')
		if cards_P_no:
			myfile.write('P set completion: ' + str(round(set_P_completion, 1)) + '%' + '\n')
		myfile.write('Foils number: ' + str(foils_no) + '\n\n')

		if missing_C:
			myfile.write('Total number of missing C: ' + str(missing_C) + '\n')
		if missing_U:
			myfile.write('Total number of missing U: ' + str(missing_U) + '\n')
		if missing_R:
			myfile.write('Total number of missing R: ' + str(missing_R) + '\n')
		if missing_P:
			myfile.write('Total number of missing P: ' + str(missing_P) + '\n')
		if missing_Rplus:
			myfile.write('Total number of missing R+: ' + str(missing_Rplus) + '\n')
		if missing_S:
			myfile.write('Total number of missing S: ' + str(missing_S) + '\n')
		if missing_RF:
			myfile.write('Total number of missing RF: ' + str(missing_RF) + '\n')
		if missing:
			myfile.write('Total number of missing cards: ' + str(missing) + '\n')

		if printMissingCards:
			myfile.write('\n=== MISSING CARDS LIST ===\n')
			if missing_C:
				myfile.write('\n')
				myfile.write('\n'.join(missing_C_log))
				myfile.write('\n')
			if missing_U:
				myfile.write('\n')
				myfile.write('\n'.join(missing_U_log))
				myfile.write('\n')
			if missing_R:
				myfile.write('\n')
				myfile.write('\n'.join(missing_R_log))
				myfile.write('\n')
			if missing_P:
				myfile.write('\n')
				myfile.write('\n'.join(missing_P_log))
				myfile.write('\n')
			if missing_Rplus:
				myfile.write('\n')
				myfile.write('\n'.join(missing_P_log))
				myfile.write('\n')
			if missing_S:
				myfile.write('\n')
				myfile.write('\n'.join(missing_P_log))
				myfile.write('\n')
			if missing_RF:
				myfile.write('\n')
				myfile.write('\n'.join(missing_P_log))
				myfile.write('\n')

from fpdf import FPDF

pdf = FPDF()
pdf.add_page()
pdf.set_font("Arial", size = 10)

f = open("lotrtcg_sets.txt", encoding='utf-8')
for x in f:
	pdf.cell(10, 4, txt = x.encode('latin-1', 'ignore').decode(), ln = 1, align = 'L')

pdf.output("lotrtcg_sets.pdf")

print('CollectionAnalyser finished - data rendered to lotrtcg_sets.txt and lotrtcg_sets.pdf')
